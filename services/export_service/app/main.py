"""
Export/Integration Service -- FastAPI on port 8015 + background stream consumer.

REST Endpoints:
  POST   /exports                 - create export job
  GET    /exports?project_id=X    - list export jobs
  GET    /exports/{id}/download   - download completed file
  GET    /exports/{id}/status     - check export status
  POST   /integrations            - configure integration
  GET    /integrations?org_id=X   - list integrations
  POST   /integrations/{id}/sync  - trigger manual sync
  DELETE /integrations/{id}       - remove integration

Background Consumer:
  Listens to STREAM_EXPORT ('export:request') and generates CSV, Excel, JSON, or PDF files.

Integration Sync:
  Pushes mention data to Salesforce, HubSpot, Slack, Tableau, or arbitrary webhooks.
"""

import asyncio
import csv
import io
import json
import logging
import os
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import httpx
from fastapi import FastAPI, HTTPException, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from sqlalchemy import select

from shared.database import create_db, init_tables
from shared.events import (
    STREAM_EXPORT,
    EventBus,
    ExportRequestEvent,
)
from shared.models import (
    ExportFormat,
    ExportJob,
    ExportStatus,
    Integration,
    Mention,
)
from shared.tracing import setup_tracing

setup_tracing("export")

logging.basicConfig(level=os.getenv("LOG_LEVEL", "INFO"))
logger = logging.getLogger(__name__)

DATABASE_URL = os.getenv("DATABASE_URL", "postgresql+asyncpg://khushfus:khushfus_dev@postgres:5432/khushfus")
REDIS_URL = os.getenv("REDIS_URL", "redis://redis:6379/0")

GROUP_NAME = "export-service"
CONSUMER_NAME = f"export-{os.getpid()}"
OUTPUT_DIR = Path("/app/exports_output")


# ============================================================
# Pydantic Schemas
# ============================================================


class ExportFilters(BaseModel):
    platform: Optional[str] = None
    sentiment: Optional[str] = None
    date_from: Optional[str] = None  # ISO date string
    date_to: Optional[str] = None  # ISO date string
    keyword: Optional[str] = None
    author_handle: Optional[str] = None
    min_likes: Optional[int] = None
    min_shares: Optional[int] = None
    is_flagged: Optional[bool] = None
    language: Optional[str] = None


class ExportCreate(BaseModel):
    project_id: int
    user_id: int = 1
    format: str = "csv"  # csv | excel | json | pdf
    filters: ExportFilters = ExportFilters()


class ExportJobOut(BaseModel):
    id: int
    project_id: int
    export_format: str
    status: str
    file_path: Optional[str] = None
    row_count: Optional[int] = None
    error_message: Optional[str] = None
    created_at: datetime
    completed_at: Optional[datetime] = None
    model_config = {"from_attributes": True}


class IntegrationCreate(BaseModel):
    organization_id: int
    integration_type: str  # salesforce | hubspot | slack | tableau | webhook
    name: str
    config: dict  # connection config (API keys, URLs, etc.)


class IntegrationOut(BaseModel):
    id: int
    organization_id: int
    integration_type: str
    name: str
    is_active: bool
    last_sync_at: Optional[datetime] = None
    created_at: datetime
    model_config = {"from_attributes": True}


class IntegrationSyncResult(BaseModel):
    integration_id: int
    status: str
    records_synced: int = 0
    error: Optional[str] = None


# ============================================================
# Export Generation Logic
# ============================================================


def _build_mention_filters(filters_json: str) -> list:
    """Parse filters JSON and return a list of SQLAlchemy where-clauses."""
    clauses = []
    if not filters_json:
        return clauses

    try:
        f = json.loads(filters_json)
    except json.JSONDecodeError:
        return clauses

    if f.get("platform"):
        clauses.append(Mention.platform == f["platform"])
    if f.get("sentiment"):
        clauses.append(Mention.sentiment == f["sentiment"])
    if f.get("date_from"):
        try:
            clauses.append(Mention.published_at >= datetime.fromisoformat(f["date_from"]))
        except ValueError:
            pass
    if f.get("date_to"):
        try:
            clauses.append(Mention.published_at <= datetime.fromisoformat(f["date_to"]))
        except ValueError:
            pass
    if f.get("keyword"):
        clauses.append(Mention.matched_keywords.ilike(f"%{f['keyword']}%"))
    if f.get("author_handle"):
        clauses.append(Mention.author_handle == f["author_handle"])
    if f.get("min_likes"):
        clauses.append(Mention.likes >= int(f["min_likes"]))
    if f.get("min_shares"):
        clauses.append(Mention.shares >= int(f["min_shares"]))
    if f.get("is_flagged") is not None:
        val = f["is_flagged"]
        if isinstance(val, str):
            val = val.lower() in ("true", "1", "yes")
        clauses.append(Mention.is_flagged.is_(bool(val)))
    if f.get("language"):
        clauses.append(Mention.language == f["language"])

    return clauses


MENTION_COLUMNS = [
    "id",
    "platform",
    "source_url",
    "text",
    "author_name",
    "author_handle",
    "author_followers",
    "likes",
    "shares",
    "comments",
    "reach",
    "sentiment",
    "sentiment_score",
    "language",
    "matched_keywords",
    "topics",
    "author_influence_score",
    "author_is_bot",
    "author_org",
    "virality_score",
    "published_at",
    "collected_at",
    "is_flagged",
]


def _mention_to_row(m: Mention) -> dict:
    """Convert a Mention ORM object to a flat dict for export."""
    return {
        "id": m.id,
        "platform": m.platform.value if m.platform else "",
        "source_url": m.source_url or "",
        "text": (m.text or "").replace("\n", " ").replace("\r", ""),
        "author_name": m.author_name or "",
        "author_handle": m.author_handle or "",
        "author_followers": m.author_followers or 0,
        "likes": m.likes,
        "shares": m.shares,
        "comments": m.comments,
        "reach": m.reach,
        "sentiment": m.sentiment.value if m.sentiment else "",
        "sentiment_score": m.sentiment_score,
        "language": m.language or "",
        "matched_keywords": m.matched_keywords or "",
        "topics": m.topics or "",
        "author_influence_score": m.author_influence_score or 0.0,
        "author_is_bot": m.author_is_bot if m.author_is_bot is not None else "",
        "author_org": m.author_org or "",
        "virality_score": m.virality_score or 0.0,
        "published_at": m.published_at.isoformat() if m.published_at else "",
        "collected_at": m.collected_at.isoformat() if m.collected_at else "",
        "is_flagged": m.is_flagged,
    }


async def _fetch_mentions(session_factory, project_id: int, filters_json: str) -> list[Mention]:
    """Fetch all mentions matching the project + filters."""
    async with session_factory() as db:
        clauses = [Mention.project_id == project_id]
        clauses.extend(_build_mention_filters(filters_json))
        result = await db.execute(select(Mention).where(*clauses).order_by(Mention.published_at.desc()))
        return result.scalars().all()


async def generate_csv_export(session_factory, job: ExportJob) -> tuple[str, int]:
    """Generate CSV file. Returns (file_path, row_count)."""
    mentions = await _fetch_mentions(session_factory, job.project_id, job.filters_json)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    filename = f"export_{job.id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
    filepath = OUTPUT_DIR / filename

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=MENTION_COLUMNS)
        writer.writeheader()
        for m in mentions:
            writer.writerow(_mention_to_row(m))

    return str(filepath), len(mentions)


async def generate_excel_export(session_factory, job: ExportJob) -> tuple[str, int]:
    """Generate Excel workbook with multiple sheets: Mentions, Sentiment Summary, Top Contributors."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    mentions = await _fetch_mentions(session_factory, job.project_id, job.filters_json)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    filename = f"export_{job.id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = OUTPUT_DIR / filename

    wb = Workbook()

    # --- Sheet 1: Mentions ---
    ws_mentions = wb.active
    ws_mentions.title = "Mentions"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, col_name in enumerate(MENTION_COLUMNS, 1):
        cell = ws_mentions.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Write data
    for row_idx, m in enumerate(mentions, 2):
        row_data = _mention_to_row(m)
        for col_idx, col_name in enumerate(MENTION_COLUMNS, 1):
            cell = ws_mentions.cell(row=row_idx, column=col_idx, value=row_data[col_name])
            cell.border = thin_border

    # Auto-width columns (approximate)
    for col_idx, col_name in enumerate(MENTION_COLUMNS, 1):
        max_len = len(col_name)
        for row_idx in range(2, min(len(mentions) + 2, 52)):  # sample first 50 rows
            val = ws_mentions.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 50))
        ws_mentions.column_dimensions[ws_mentions.cell(row=1, column=col_idx).column_letter].width = max_len + 2

    # --- Sheet 2: Sentiment Summary ---
    ws_sentiment = wb.create_sheet("Sentiment Summary")

    sentiment_counts = {}
    for m in mentions:
        s = m.sentiment.value if m.sentiment else "unknown"
        sentiment_counts[s] = sentiment_counts.get(s, 0) + 1

    ws_sentiment.cell(row=1, column=1, value="Sentiment").font = header_font
    ws_sentiment.cell(row=1, column=1).fill = header_fill
    ws_sentiment.cell(row=1, column=2, value="Count").font = header_font
    ws_sentiment.cell(row=1, column=2).fill = header_fill
    ws_sentiment.cell(row=1, column=3, value="Percentage").font = header_font
    ws_sentiment.cell(row=1, column=3).fill = header_fill

    total = len(mentions) or 1
    for row_idx, (sentiment, count) in enumerate(sorted(sentiment_counts.items()), 2):
        ws_sentiment.cell(row=row_idx, column=1, value=sentiment.capitalize())
        ws_sentiment.cell(row=row_idx, column=2, value=count)
        ws_sentiment.cell(row=row_idx, column=3, value=f"{count / total * 100:.1f}%")

    # Summary row
    summary_row = len(sentiment_counts) + 3
    ws_sentiment.cell(row=summary_row, column=1, value="Total Mentions").font = Font(bold=True)
    ws_sentiment.cell(row=summary_row, column=2, value=len(mentions))

    avg_score = sum(m.sentiment_score for m in mentions) / total if mentions else 0
    ws_sentiment.cell(row=summary_row + 1, column=1, value="Average Sentiment Score").font = Font(bold=True)
    ws_sentiment.cell(row=summary_row + 1, column=2, value=round(avg_score, 4))

    ws_sentiment.column_dimensions["A"].width = 25
    ws_sentiment.column_dimensions["B"].width = 12
    ws_sentiment.column_dimensions["C"].width = 14

    # --- Sheet 3: Top Contributors ---
    ws_contributors = wb.create_sheet("Top Contributors")

    # Aggregate by author
    author_stats: dict[str, dict] = {}
    for m in mentions:
        handle = m.author_handle or "unknown"
        if handle not in author_stats:
            author_stats[handle] = {
                "name": m.author_name or "",
                "handle": handle,
                "followers": m.author_followers or 0,
                "platform": m.platform.value if m.platform else "",
                "mentions": 0,
                "total_likes": 0,
                "total_shares": 0,
                "influence": m.author_influence_score or 0.0,
            }
        author_stats[handle]["mentions"] += 1
        author_stats[handle]["total_likes"] += m.likes
        author_stats[handle]["total_shares"] += m.shares
        # Keep the highest follower count
        if (m.author_followers or 0) > author_stats[handle]["followers"]:
            author_stats[handle]["followers"] = m.author_followers or 0

    # Sort by mention count descending, take top 50
    top_authors = sorted(author_stats.values(), key=lambda x: x["mentions"], reverse=True)[:50]

    contrib_columns = [
        "Name",
        "Handle",
        "Platform",
        "Followers",
        "Mentions",
        "Total Likes",
        "Total Shares",
        "Influence Score",
    ]
    for col_idx, col_name in enumerate(contrib_columns, 1):
        cell = ws_contributors.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, author in enumerate(top_authors, 2):
        ws_contributors.cell(row=row_idx, column=1, value=author["name"])
        ws_contributors.cell(row=row_idx, column=2, value=author["handle"])
        ws_contributors.cell(row=row_idx, column=3, value=author["platform"])
        ws_contributors.cell(row=row_idx, column=4, value=author["followers"])
        ws_contributors.cell(row=row_idx, column=5, value=author["mentions"])
        ws_contributors.cell(row=row_idx, column=6, value=author["total_likes"])
        ws_contributors.cell(row=row_idx, column=7, value=author["total_shares"])
        ws_contributors.cell(row=row_idx, column=8, value=round(author["influence"], 2))

    for col_idx in range(1, len(contrib_columns) + 1):
        ws_contributors.column_dimensions[ws_contributors.cell(row=1, column=col_idx).column_letter].width = 18

    wb.save(str(filepath))
    return str(filepath), len(mentions)


async def generate_json_export(session_factory, job: ExportJob) -> tuple[str, int]:
    """Generate JSON file with filtered mentions."""
    mentions = await _fetch_mentions(session_factory, job.project_id, job.filters_json)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    filename = f"export_{job.id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
    filepath = OUTPUT_DIR / filename

    rows = [_mention_to_row(m) for m in mentions]
    export_data = {
        "export_id": job.id,
        "project_id": job.project_id,
        "generated_at": datetime.utcnow().isoformat(),
        "total_records": len(rows),
        "mentions": rows,
    }

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(export_data, f, indent=2, default=str)

    return str(filepath), len(mentions)


async def generate_pdf_export(session_factory, job: ExportJob) -> tuple[str, int]:
    """
    Generate PDF export reusing report template rendering approach.
    Falls back to HTML if WeasyPrint is not available.
    """
    from jinja2 import BaseLoader, Environment

    mentions = await _fetch_mentions(session_factory, job.project_id, job.filters_json)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Build summary data
    sentiment_counts = {}
    for m in mentions:
        s = m.sentiment.value if m.sentiment else "unknown"
        sentiment_counts[s] = sentiment_counts.get(s, 0) + 1

    total = len(mentions) or 1
    avg_score = sum(m.sentiment_score for m in mentions) / total if mentions else 0

    platform_counts = {}
    for m in mentions:
        p = m.platform.value if m.platform else "unknown"
        platform_counts[p] = platform_counts.get(p, 0) + 1

    # Top contributors
    author_stats: dict[str, dict] = {}
    for m in mentions:
        handle = m.author_handle or "unknown"
        if handle not in author_stats:
            author_stats[handle] = {
                "name": m.author_name or "",
                "handle": handle,
                "followers": m.author_followers or 0,
                "mentions": 0,
            }
        author_stats[handle]["mentions"] += 1
    top_authors = sorted(author_stats.values(), key=lambda x: x["mentions"], reverse=True)[:20]

    template_str = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>KhushFus Export Report</title>
        <style>
            body { font-family: 'Segoe UI', Arial, sans-serif; margin: 40px; color: #333; }
            h1 { color: #2F5496; border-bottom: 3px solid #2F5496; padding-bottom: 10px; }
            h2 { color: #2F5496; margin-top: 30px; }
            table { border-collapse: collapse; width: 100%; margin: 15px 0; }
            th { background-color: #2F5496; color: white; padding: 10px 12px; text-align: left; }
            td { padding: 8px 12px; border-bottom: 1px solid #ddd; }
            tr:nth-child(even) { background-color: #f8f9fa; }
            .summary-box { background: #f0f4f8; border-radius: 8px; padding: 20px; margin: 20px 0; }
            .summary-box .stat { display: inline-block; margin-right: 40px; }
            .summary-box .stat-value { font-size: 24px; font-weight: bold; color: #2F5496; }
            .summary-box .stat-label { font-size: 12px; color: #666; }
            .footer { margin-top: 40px; text-align: center; font-size: 11px; color: #999; }
        </style>
    </head>
    <body>
        <h1>KhushFus Data Export</h1>
        <p>Generated: {{ generated_at }} | Project ID: {{ project_id }} | Total Mentions: {{ total }}</p>

        <div class="summary-box">
            <div class="stat">
                <div class="stat-value">{{ total }}</div>
                <div class="stat-label">Total Mentions</div>
            </div>
            <div class="stat">
                <div class="stat-value">{{ "%.2f"|format(avg_score) }}</div>
                <div class="stat-label">Avg Sentiment Score</div>
            </div>
            {% for s, c in sentiment_counts.items() %}
            <div class="stat">
                <div class="stat-value">{{ c }}</div>
                <div class="stat-label">{{ s|capitalize }}</div>
            </div>
            {% endfor %}
        </div>

        <h2>Platform Breakdown</h2>
        <table>
            <tr><th>Platform</th><th>Count</th><th>Percentage</th></tr>
            {% for p, c in platform_counts.items() %}
            <tr><td>{{ p|capitalize }}</td><td>{{ c }}</td><td>{{ "%.1f"|format(c / total * 100) }}%</td></tr>
            {% endfor %}
        </table>

        <h2>Top Contributors</h2>
        <table>
            <tr><th>Name</th><th>Handle</th><th>Followers</th><th>Mentions</th></tr>
            {% for a in top_authors %}
            <tr><td>{{ a.name }}</td><td>{{ a.handle }}</td><td>{{ a.followers }}</td><td>{{ a.mentions }}</td></tr>
            {% endfor %}
        </table>

        <h2>Mentions (first 100)</h2>
        <table>
            <tr><th>Author</th><th>Platform</th><th>Text</th><th>Sentiment</th><th>Likes</th><th>Shares</th></tr>
            {% for m in mention_rows[:100] %}
            <tr>
                <td>{{ m.author_handle }}</td>
                <td>{{ m.platform }}</td>
                <td>{{ m.text[:150] }}{% if m.text|length > 150 %}...{% endif %}</td>
                <td>{{ m.sentiment }}</td>
                <td>{{ m.likes }}</td>
                <td>{{ m.shares }}</td>
            </tr>
            {% endfor %}
        </table>
        {% if total > 100 %}
        <p><em>Showing 100 of {{ total }} mentions. Use CSV or Excel export for full data.</em></p>
        {% endif %}

        <div class="footer">
            KhushFus Social Listening Platform &mdash; Confidential
        </div>
    </body>
    </html>
    """

    env = Environment(loader=BaseLoader())
    template = env.from_string(template_str)
    mention_rows = [_mention_to_row(m) for m in mentions]

    html = template.render(
        generated_at=datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
        project_id=job.project_id,
        total=total,
        avg_score=avg_score,
        sentiment_counts=sentiment_counts,
        platform_counts=platform_counts,
        top_authors=top_authors,
        mention_rows=mention_rows,
    )

    # Try PDF via WeasyPrint, fall back to HTML
    filename_base = f"export_{job.id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
    try:
        from weasyprint import HTML as WeasyprintHTML  # noqa: N811

        filepath = OUTPUT_DIR / f"{filename_base}.pdf"
        WeasyprintHTML(string=html).write_pdf(str(filepath))
    except ImportError:
        logger.warning("WeasyPrint not available, generating HTML instead of PDF")
        filepath = OUTPUT_DIR / f"{filename_base}.html"
        filepath.write_text(html, encoding="utf-8")

    return str(filepath), len(mentions)


EXPORT_GENERATORS = {
    ExportFormat.CSV: generate_csv_export,
    ExportFormat.EXCEL: generate_excel_export,
    ExportFormat.JSON: generate_json_export,
    ExportFormat.PDF: generate_pdf_export,
}


async def process_export_job(session_factory, job_id: int):
    """Execute an export job: set status, generate file, update record."""
    async with session_factory() as db:
        job = await db.get(ExportJob, job_id)
        if not job:
            logger.error(f"Export job {job_id} not found")
            return

        job.status = ExportStatus.PROCESSING
        await db.commit()

    try:
        fmt_val = job.export_format.value if hasattr(job.export_format, "value") else job.export_format
        generator = EXPORT_GENERATORS.get(ExportFormat(fmt_val))
        if not generator:
            raise ValueError(f"Unsupported export format: {job.export_format}")

        file_path, row_count = await generator(session_factory, job)

        async with session_factory() as db:
            job = await db.get(ExportJob, job_id)
            job.status = ExportStatus.COMPLETED
            job.file_path = file_path
            job.row_count = row_count
            job.completed_at = datetime.utcnow()
            await db.commit()

        logger.info(f"Export job {job_id} completed: {row_count} rows -> {file_path}")

    except Exception as e:
        logger.error(f"Export job {job_id} failed: {e}", exc_info=True)
        async with session_factory() as db:
            job = await db.get(ExportJob, job_id)
            if job:
                job.status = ExportStatus.FAILED
                job.error_message = str(e)[:500]
                await db.commit()


# ============================================================
# Integration Sync Logic
# ============================================================


async def sync_integration(
    session_factory,
    integration_id: int,
    project_id: int | None = None,
) -> IntegrationSyncResult:
    """Push mention data to the configured external system."""
    async with session_factory() as db:
        integration = await db.get(Integration, integration_id)
        if not integration:
            return IntegrationSyncResult(integration_id=integration_id, status="error", error="Integration not found")

        if not integration.is_active:
            return IntegrationSyncResult(integration_id=integration_id, status="error", error="Integration is inactive")

        try:
            config = json.loads(integration.config_json)
        except json.JSONDecodeError:
            return IntegrationSyncResult(integration_id=integration_id, status="error", error="Invalid configuration")

        # Determine the project to sync (from config or param)
        sync_project_id = project_id or config.get("project_id")
        if not sync_project_id:
            return IntegrationSyncResult(
                integration_id=integration_id, status="error", error="No project_id configured"
            )

        # Fetch recent mentions (since last sync or last 24h)
        since = integration.last_sync_at or (datetime.utcnow() - timedelta(hours=24))
        result = await db.execute(
            select(Mention)
            .where(
                Mention.project_id == int(sync_project_id),
                Mention.collected_at >= since,
            )
            .order_by(Mention.collected_at.desc())
            .limit(1000)
        )
        mentions = result.scalars().all()

        if not mentions:
            integration.last_sync_at = datetime.utcnow()
            await db.commit()
            return IntegrationSyncResult(integration_id=integration_id, status="success", records_synced=0)

        mention_data = [_mention_to_row(m) for m in mentions]

    # Dispatch to integration-specific handler
    handler = INTEGRATION_HANDLERS.get(integration.integration_type)
    if not handler:
        return IntegrationSyncResult(
            integration_id=integration_id,
            status="error",
            error=f"Unknown integration type: {integration.integration_type}",
        )

    try:
        synced = await handler(config, mention_data)

        async with session_factory() as db:
            integration = await db.get(Integration, integration_id)
            integration.last_sync_at = datetime.utcnow()
            await db.commit()

        return IntegrationSyncResult(integration_id=integration_id, status="success", records_synced=synced)

    except Exception as e:
        logger.error(f"Integration sync failed for {integration_id}: {e}", exc_info=True)
        return IntegrationSyncResult(integration_id=integration_id, status="error", error=str(e)[:500])


async def _sync_webhook(config: dict, mentions: list[dict]) -> int:
    """Push mentions to a webhook URL as JSON payload."""
    url = config.get("webhook_url")
    if not url:
        raise ValueError("webhook_url not configured")

    headers = config.get("headers", {})
    secret = config.get("secret")
    if secret:
        headers["X-Webhook-Secret"] = secret

    async with httpx.AsyncClient(timeout=30.0) as client:
        # Send in batches of 100
        synced = 0
        for i in range(0, len(mentions), 100):
            batch = mentions[i : i + 100]
            resp = await client.post(
                url,
                json={"mentions": batch, "count": len(batch), "timestamp": datetime.utcnow().isoformat()},
                headers=headers,
            )
            resp.raise_for_status()
            synced += len(batch)

    return synced


async def _sync_slack(config: dict, mentions: list[dict]) -> int:
    """Post a summary to a Slack webhook."""
    webhook_url = config.get("webhook_url")
    if not webhook_url:
        raise ValueError("Slack webhook_url not configured")

    channel = config.get("channel", "#social-listening")
    total = len(mentions)

    # Build sentiment breakdown
    sentiment_counts: dict[str, int] = {}
    for m in mentions:
        s = m.get("sentiment", "unknown")
        sentiment_counts[s] = sentiment_counts.get(s, 0) + 1

    sentiment_text = " | ".join(f"{k.capitalize()}: {v}" for k, v in sorted(sentiment_counts.items()))

    # Top mentions by engagement
    top = sorted(mentions, key=lambda x: x.get("likes", 0) + x.get("shares", 0), reverse=True)[:5]
    top_text = "\n".join(
        f"  - @{m.get('author_handle', '?')} ({m.get('platform', '?')}): "
        f"{m.get('text', '')[:100]}... [{m.get('likes', 0)} likes, {m.get('shares', 0)} shares]"
        for m in top
    )

    blocks = [
        {"type": "header", "text": {"type": "plain_text", "text": f"KhushFus Sync: {total} new mentions"}},
        {"type": "section", "text": {"type": "mrkdwn", "text": f"*Sentiment:* {sentiment_text}"}},
        {"type": "section", "text": {"type": "mrkdwn", "text": f"*Top Mentions:*\n{top_text}"}},
    ]

    async with httpx.AsyncClient(timeout=15.0) as client:
        resp = await client.post(webhook_url, json={"channel": channel, "blocks": blocks})
        resp.raise_for_status()

    return total


async def _sync_salesforce(config: dict, mentions: list[dict]) -> int:
    """Push mentions to Salesforce as custom object records via REST API."""
    instance_url = config.get("instance_url")
    access_token = config.get("access_token")
    object_name = config.get("object_name", "Social_Mention__c")

    if not instance_url or not access_token:
        raise ValueError("Salesforce instance_url and access_token required")

    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json",
    }

    async with httpx.AsyncClient(timeout=30.0) as client:
        synced = 0
        for m in mentions:
            payload = {
                "Platform__c": m.get("platform", ""),
                "Author_Handle__c": m.get("author_handle", "")[:80],
                "Author_Name__c": m.get("author_name", "")[:255],
                "Text__c": m.get("text", "")[:32000],
                "Sentiment__c": m.get("sentiment", ""),
                "Sentiment_Score__c": m.get("sentiment_score", 0),
                "Likes__c": m.get("likes", 0),
                "Shares__c": m.get("shares", 0),
                "Source_URL__c": m.get("source_url", "")[:255],
                "Published_At__c": m.get("published_at", ""),
            }
            try:
                resp = await client.post(
                    f"{instance_url}/services/data/v58.0/sobjects/{object_name}/",
                    json=payload,
                    headers=headers,
                )
                if resp.status_code in (200, 201):
                    synced += 1
                else:
                    logger.warning(f"Salesforce create failed: {resp.status_code} {resp.text[:200]}")
            except Exception as e:
                logger.warning(f"Salesforce record sync error: {e}")

    return synced


async def _sync_hubspot(config: dict, mentions: list[dict]) -> int:
    """Push mentions to HubSpot as custom object records or engagements."""
    api_key = config.get("api_key")
    if not api_key:
        raise ValueError("HubSpot api_key required")

    base_url = "https://api.hubapi.com"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }

    async with httpx.AsyncClient(timeout=30.0) as client:
        synced = 0
        for m in mentions:
            # Create as engagement/note
            payload = {
                "properties": {
                    "hs_timestamp": m.get("published_at") or datetime.utcnow().isoformat(),
                    "hs_note_body": (
                        f"[{m.get('platform', 'social')}] @{m.get('author_handle', '?')}: "
                        f"{m.get('text', '')[:500]}\n\n"
                        f"Sentiment: {m.get('sentiment', 'N/A')} ({m.get('sentiment_score', 0)})\n"
                        f"Likes: {m.get('likes', 0)} | Shares: {m.get('shares', 0)}\n"
                        f"URL: {m.get('source_url', '')}"
                    ),
                }
            }
            try:
                resp = await client.post(
                    f"{base_url}/crm/v3/objects/notes",
                    json=payload,
                    headers=headers,
                )
                if resp.status_code in (200, 201):
                    synced += 1
                else:
                    logger.warning(f"HubSpot create failed: {resp.status_code}")
            except Exception as e:
                logger.warning(f"HubSpot record sync error: {e}")

    return synced


async def _sync_tableau(config: dict, mentions: list[dict]) -> int:
    """Push data to Tableau via Hyper API or REST (publish as data source)."""
    server_url = config.get("server_url")
    token_name = config.get("token_name")
    token_value = config.get("token_value")
    site_id = config.get("site_id", "")

    if not server_url or not token_name or not token_value:
        raise ValueError("Tableau server_url, token_name, and token_value required")

    # Authenticate
    async with httpx.AsyncClient(timeout=30.0) as client:
        auth_resp = await client.post(
            f"{server_url}/api/3.19/auth/signin",
            json={
                "credentials": {
                    "personalAccessTokenName": token_name,
                    "personalAccessTokenSecret": token_value,
                    "site": {"contentUrl": site_id},
                }
            },
        )
        auth_resp.raise_for_status()
        auth_data = auth_resp.json()
        credentials = auth_data.get("credentials", {})
        api_token = credentials.get("token", "")
        site_api_id = credentials.get("site", {}).get("id", "")

        if not api_token:
            raise ValueError("Tableau authentication failed")

        # For Tableau, write the data as a CSV file and publish as data source
        # This is a simplified approach; production would use Hyper API
        csv_buffer = io.StringIO()
        if mentions:
            writer = csv.DictWriter(csv_buffer, fieldnames=list(mentions[0].keys()))
            writer.writeheader()
            for m in mentions:
                writer.writerow(m)

        csv_content = csv_buffer.getvalue().encode("utf-8")

        # Publish data source
        headers = {
            "X-Tableau-Auth": api_token,
            "Content-Type": "application/octet-stream",
        }
        publish_resp = await client.post(
            f"{server_url}/api/3.19/sites/{site_api_id}/datasources",
            params={"datasourceName": f"khushfus_mentions_{datetime.utcnow().strftime('%Y%m%d')}", "overwrite": "true"},
            content=csv_content,
            headers=headers,
        )

        if publish_resp.status_code in (200, 201):
            return len(mentions)
        else:
            logger.warning(f"Tableau publish failed: {publish_resp.status_code}")
            return 0


INTEGRATION_HANDLERS = {
    "webhook": _sync_webhook,
    "slack": _sync_slack,
    "salesforce": _sync_salesforce,
    "hubspot": _sync_hubspot,
    "tableau": _sync_tableau,
}


# ============================================================
# Background Consumer
# ============================================================


async def export_consumer_loop(bus: EventBus, session_factory):
    """Background consumer: listen for export requests on the stream."""
    await bus.ensure_group(STREAM_EXPORT, GROUP_NAME)
    logger.info("Export consumer listening on '%s'...", STREAM_EXPORT)

    while True:
        try:
            messages = await bus.consume(
                STREAM_EXPORT,
                GROUP_NAME,
                CONSUMER_NAME,
                count=5,
                block_ms=3000,
            )

            for msg_id, data in messages:
                try:
                    job_id = int(data.get("export_job_id", 0))
                    if job_id == 0:
                        logger.warning(f"Invalid export_job_id in request: {data}")
                        continue

                    await process_export_job(session_factory, job_id)

                except Exception as e:
                    logger.error(f"Export consumer error for {msg_id}: {e}", exc_info=True)
                finally:
                    await bus.ack(STREAM_EXPORT, GROUP_NAME, msg_id)

        except Exception as e:
            logger.error(f"Export consumer loop error: {e}", exc_info=True)
            await asyncio.sleep(1)


# ============================================================
# FastAPI Application
# ============================================================


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Start DB, EventBus, and background consumer on startup; clean up on shutdown."""
    engine, session_factory = create_db(DATABASE_URL)
    await init_tables(engine)

    bus = EventBus(REDIS_URL)
    await bus.connect()

    app.state.db_session = session_factory
    app.state.event_bus = bus

    # Start background consumer as a task
    consumer_task = asyncio.create_task(export_consumer_loop(bus, session_factory))

    yield

    consumer_task.cancel()
    try:
        await consumer_task
    except asyncio.CancelledError:
        pass
    await bus.close()
    await engine.dispose()


app = FastAPI(
    title="KhushFus Export & Integration Service",
    description="Export data and sync with external platforms",
    version="0.1.0",
    contact={"name": "KhushFus Engineering", "email": "engineering@khushfus.io"},
    license_info={"name": "Proprietary"},
    openapi_tags=[
        {"name": "Exports", "description": "Create, list, check status, and download export jobs."},
        {"name": "Integrations", "description": "Configure and manage external platform integrations."},
        {"name": "Health", "description": "Service health check."},
    ],
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


try:
    from prometheus_fastapi_instrumentator import Instrumentator

    Instrumentator().instrument(app).expose(app)
except ImportError:
    pass


# ============================================================
# Health Check
# ============================================================


@app.get(
    "/health",
    tags=["Health"],
    summary="Export health check",
    description="Returns the health status of the Export & Integration service and its dependencies.",
)
async def health():
    from shared.health import build_health_response, check_postgres, check_redis

    checks = {
        "postgres": await check_postgres(database_url=DATABASE_URL),
        "redis": await check_redis(REDIS_URL),
    }
    return await build_health_response("export-service", checks=checks)


# ============================================================
# Export Endpoints
# ============================================================


@app.post(
    "/exports",
    response_model=ExportJobOut,
    status_code=201,
    tags=["Exports"],
    summary="Create an export job",
    description="Create a new export job. Supports CSV, Excel, JSON, and PDF formats with filters.",
)
async def create_export(payload: ExportCreate, request: Request):
    """Create a new export job. Publishes to the export stream for async processing."""
    session_factory = request.app.state.db_session
    bus: EventBus = request.app.state.event_bus

    # Validate format
    try:
        fmt = ExportFormat(payload.format)
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid format: {payload.format}. Must be csv, excel, json, or pdf.",
        )

    filters_json = payload.filters.model_dump_json() if payload.filters else ""

    async with session_factory() as db:
        job = ExportJob(
            project_id=payload.project_id,
            user_id=payload.user_id,
            export_format=fmt,
            filters_json=filters_json,
            status=ExportStatus.PENDING,
        )
        db.add(job)
        await db.commit()
        await db.refresh(job)

        # Publish to stream for background processing
        await bus.publish(
            STREAM_EXPORT,
            ExportRequestEvent(
                export_job_id=job.id,
                project_id=job.project_id,
                export_format=fmt.value,
                filters_json=filters_json,
            ),
        )

        return ExportJobOut(
            id=job.id,
            project_id=job.project_id,
            export_format=fmt.value,
            status=job.status.value,
            file_path=job.file_path,
            row_count=job.row_count,
            error_message=job.error_message,
            created_at=job.created_at,
            completed_at=job.completed_at,
        )


@app.get(
    "/exports",
    response_model=list[ExportJobOut],
    tags=["Exports"],
    summary="List export jobs",
    description="List all export jobs for a project, sorted by creation date descending.",
)
async def list_exports(request: Request, project_id: int = Query(...)):
    """List export jobs for a project."""
    session_factory = request.app.state.db_session

    async with session_factory() as db:
        result = await db.execute(
            select(ExportJob).where(ExportJob.project_id == project_id).order_by(ExportJob.created_at.desc())
        )
        jobs = result.scalars().all()

        return [
            ExportJobOut(
                id=j.id,
                project_id=j.project_id,
                export_format=j.export_format.value if hasattr(j.export_format, "value") else j.export_format,
                status=j.status.value if hasattr(j.status, "value") else j.status,
                file_path=j.file_path,
                row_count=j.row_count,
                error_message=j.error_message,
                created_at=j.created_at,
                completed_at=j.completed_at,
            )
            for j in jobs
        ]


@app.get(
    "/exports/{export_id}/status",
    tags=["Exports"],
    summary="Check export status",
    description="Check the current status of an export job including row count and any error messages.",
)
async def get_export_status(export_id: int, request: Request):
    """Check the status of an export job."""
    session_factory = request.app.state.db_session

    async with session_factory() as db:
        job = await db.get(ExportJob, export_id)
        if not job:
            raise HTTPException(status_code=404, detail="Export job not found")

        return {
            "id": job.id,
            "status": job.status.value if hasattr(job.status, "value") else job.status,
            "row_count": job.row_count,
            "file_path": job.file_path,
            "error_message": job.error_message,
            "created_at": job.created_at.isoformat() if job.created_at else None,
            "completed_at": job.completed_at.isoformat() if job.completed_at else None,
        }


@app.get(
    "/exports/{export_id}/download",
    tags=["Exports"],
    summary="Download export file",
    description="Download the completed export file. Returns 409 if the export is not yet complete.",
)
async def download_export(export_id: int, request: Request):
    """Download the completed export file."""
    session_factory = request.app.state.db_session

    async with session_factory() as db:
        job = await db.get(ExportJob, export_id)
        if not job:
            raise HTTPException(status_code=404, detail="Export job not found")

        if job.status != ExportStatus.COMPLETED:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"Export is not ready. Current status: "
                    f"{job.status.value if hasattr(job.status, 'value') else job.status}"
                ),
            )

        if not job.file_path or not Path(job.file_path).exists():
            raise HTTPException(status_code=404, detail="Export file not found on disk")

        filepath = Path(job.file_path)
        media_types = {
            ".csv": "text/csv",
            ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ".json": "application/json",
            ".pdf": "application/pdf",
            ".html": "text/html",
        }
        media_type = media_types.get(filepath.suffix, "application/octet-stream")

        return FileResponse(
            path=str(filepath),
            media_type=media_type,
            filename=filepath.name,
        )


# ============================================================
# Integration Endpoints
# ============================================================


@app.post(
    "/integrations",
    response_model=IntegrationOut,
    status_code=201,
    tags=["Integrations"],
    summary="Create an integration",
    description="Configure a new integration with Salesforce, HubSpot, Slack, Tableau, or a custom webhook.",
)
async def create_integration(payload: IntegrationCreate, request: Request):
    """Configure a new integration (Salesforce, HubSpot, Slack, Tableau, webhook)."""
    session_factory = request.app.state.db_session

    valid_types = {"salesforce", "hubspot", "slack", "tableau", "webhook"}
    if payload.integration_type not in valid_types:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid integration_type. Must be one of: {', '.join(sorted(valid_types))}",
        )

    async with session_factory() as db:
        integration = Integration(
            organization_id=payload.organization_id,
            integration_type=payload.integration_type,
            name=payload.name,
            config_json=json.dumps(payload.config),
            is_active=True,
        )
        db.add(integration)
        await db.commit()
        await db.refresh(integration)

        return IntegrationOut(
            id=integration.id,
            organization_id=integration.organization_id,
            integration_type=integration.integration_type,
            name=integration.name,
            is_active=integration.is_active,
            last_sync_at=integration.last_sync_at,
            created_at=integration.created_at,
        )


@app.get(
    "/integrations",
    response_model=list[IntegrationOut],
    tags=["Integrations"],
    summary="List integrations",
    description="List all configured integrations for an organization.",
)
async def list_integrations(request: Request, org_id: int = Query(...)):
    """List integrations for an organization."""
    session_factory = request.app.state.db_session

    async with session_factory() as db:
        result = await db.execute(
            select(Integration).where(Integration.organization_id == org_id).order_by(Integration.created_at.desc())
        )
        integrations = result.scalars().all()

        return [
            IntegrationOut(
                id=i.id,
                organization_id=i.organization_id,
                integration_type=i.integration_type,
                name=i.name,
                is_active=i.is_active,
                last_sync_at=i.last_sync_at,
                created_at=i.created_at,
            )
            for i in integrations
        ]


@app.post(
    "/integrations/{integration_id}/sync",
    response_model=IntegrationSyncResult,
    tags=["Integrations"],
    summary="Trigger integration sync",
    description="Manually trigger a sync for an integration, pushing mention data to the external platform.",
)
async def trigger_sync(integration_id: int, request: Request, project_id: int = Query(None)):
    """Trigger a manual sync for an integration."""
    session_factory = request.app.state.db_session

    result = await sync_integration(session_factory, integration_id, project_id)

    if result.status == "error":
        raise HTTPException(status_code=400, detail=result.error)

    return result


@app.delete(
    "/integrations/{integration_id}",
    status_code=204,
    tags=["Integrations"],
    summary="Delete an integration",
    description="Remove an integration configuration permanently.",
)
async def delete_integration(integration_id: int, request: Request):
    """Delete an integration."""
    session_factory = request.app.state.db_session

    async with session_factory() as db:
        integration = await db.get(Integration, integration_id)
        if not integration:
            raise HTTPException(status_code=404, detail="Integration not found")

        await db.delete(integration)
        await db.commit()

    return None
